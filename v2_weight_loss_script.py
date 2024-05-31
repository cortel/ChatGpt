import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Initial values
initial_weight = 97.0
initial_fat_percentage = 30.0
daily_distance_km = 10.00
calories_per_kg_per_km = 0.9
calories_per_kg_fat = 7700
target_fat_percentage = 6.0

# Initial calculations
initial_fat_mass = initial_weight * (initial_fat_percentage / 100)
initial_lean_body_mass = initial_weight - initial_fat_mass

# Create a list to store results
data = []

current_weight = initial_weight
current_fat_percentage = initial_fat_percentage
current_fat_mass = initial_fat_mass
current_day = 1

print("Processing daily data...")

while current_fat_percentage > target_fat_percentage:
    # Calculate calories burned daily
    calories_burned_daily = current_weight * daily_distance_km * calories_per_kg_per_km
    
    # Store data for each day
    data.append([
        current_day,
        initial_weight,
        initial_fat_percentage,
        daily_distance_km,
        calories_burned_daily,
        current_fat_mass,
        current_weight,
        current_fat_percentage
    ])
    print(f"Day {current_day}: Weight = {current_weight:.2f} kg, Fat Mass = {current_fat_mass:.2f} kg, Fat Percentage = {current_fat_percentage:.2f}%, Calories Burned = {calories_burned_daily:.2f} kcal")
    
    # Update fat mass based on calories burned
    fat_calories_burned = calories_burned_daily / calories_per_kg_fat
    current_fat_mass -= fat_calories_burned
    
    # Update weight and fat percentage
    current_weight -= calories_burned_daily / calories_per_kg_fat
    current_fat_percentage = (current_fat_mass / current_weight) * 100
    
    # Prepare for next iteration
    initial_weight = current_weight
    initial_fat_percentage = current_fat_percentage
    
    current_day += 1

# Create DataFrame for daily progress
columns = ['Day', 'Initial Weight (kg)', 'Initial Fat Percentage (%)', 'Km Walked Daily', 'Calories Burned Daily (kcal)', 'Current Fat Mass (kg)', 'Current Weight (kg)', 'Current Fat Percentage (%)']
df_daily_progress = pd.DataFrame(data, columns=columns)

# Create DataFrame for constants
constants_data = {
    'Distance Walked Daily (km)': [daily_distance_km],
    'Initial Calories Burned Daily (kcal)': [initial_weight * daily_distance_km * calories_per_kg_per_km]
}
df_constants = pd.DataFrame(constants_data)

# Create DataFrame for initial values and thought process
thought_process_data = {
    'Initial Values': ['Initial Weight', 'Initial Fat Percentage', 'Daily Distance Walked', 'Calories Burned per kg per km', 'Calories per kg of Fat', 'Target Fat Percentage'],
    'Values': [initial_weight, initial_fat_percentage, daily_distance_km, calories_per_kg_per_km, calories_per_kg_fat, target_fat_percentage],
    'Formula/Thought Process': [
        'The starting weight of the individual.',
        'The starting fat percentage of the individual.',
        'The distance the individual plans to walk daily.',
        'The average number of calories burned per kg of body weight per km walked.',
        'The amount of calories that need to be burned to lose 1 kg of fat.',
        'The target fat percentage the individual aims to reach.'
    ]
}
df_thought_process = pd.DataFrame(thought_process_data)

# Save to Excel file with multiple sheets
file_path = 'weight_loss_daily_progress_with_constants.xlsx'
with pd.ExcelWriter(file_path) as writer:
    df_daily_progress.to_excel(writer, sheet_name='Daily Progress', index=False)
    df_constants.to_excel(writer, sheet_name='Constants', index=False)
    df_thought_process.to_excel(writer, sheet_name='Initial Values and Thoughts', index=False)

# Load the workbook and select the active worksheet
wb = load_workbook(file_path)
ws = wb['Daily Progress']

# Highlight color for the separation row
fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Insert empty highlighted rows for weekly separation
for row in range(len(df_daily_progress) // 7, 0, -1):
    ws.insert_rows(idx=row*7+1)
    for cell in ws[row*7+1]:
        cell.fill = fill

# Save the updated workbook
wb.save(file_path)

print(f"\nExcel file saved to {file_path}")
